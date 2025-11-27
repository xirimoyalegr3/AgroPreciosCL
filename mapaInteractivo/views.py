# mapaInteractivo/views.py
from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.db.models import Count, Avg, Min, Max, Sum, F
from django.db import transaction
from .models import Region, DatosComercializacion, Producto, Mercado, Subsector
import logging
import io
from django.core.cache import cache
from django.views.decorators.cache import cache_page

# Imports para PDF
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch

# Imports para Excel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


def mapa_interactivo(request):
    """Vista principal del mapa interactivo"""
    try:
        return render(request, 'mapaInteractivo/mapa.html')
    except Exception as e:
        logger.error(f"Error en vista mapa_interactivo: {str(e)}")
        return render(request, 'mapaInteractivo/error.html', {'error': str(e)})

@cache_page(60 * 240)  # 10 minutos  
def obtener_datos_region(request, region_id):
    """API para obtener datos REALES de una region especifica"""
    try:
        with transaction.atomic():
            region = Region.objects.get(id_region=region_id)

            total_registros_region = DatosComercializacion.objects.filter(region=region).count()

            productos_region = (DatosComercializacion.objects
                              .filter(region=region)
                              .values('producto__nombre')
                              .distinct()
                              .count())

            mercados_region = (DatosComercializacion.objects
                              .filter(region=region)
                              .values('mercado__nombre')
                              .distinct()
                              .count())

            precios_promedio = (DatosComercializacion.objects
                               .filter(region=region)
                               .values('producto__nombre')
                               .annotate(
                                   precio_promedio=Avg('precio_promedio'),
                                   volumen_total=Sum('volumen')
                               )
                               .order_by('-volumen_total')[:10])

            subsectores = (DatosComercializacion.objects
                          .filter(region=region)
                          .values('subsector__nombre')
                          .annotate(total=Count('id'))
                          .order_by('-total'))

            subsectores_formateados = []
            for subsector in subsectores:
                subsectores_formateados.append({
                    'nombre': subsector['subsector__nombre'],
                    'total': subsector['total']
                })

            datos = {
                'region_id': region.id_region,
                'region_nombre': region.nombre,
                'total_registros': total_registros_region,
                'total_productos': productos_region,
                'total_mercados': mercados_region,
                'subsectores': subsectores_formateados,
                'productos_destacados': list(precios_promedio),
            }

            return JsonResponse(datos)

    except Region.DoesNotExist:
        logger.warning(f"Region con ID {region_id} no encontrada")
        return JsonResponse({'error': f'Region con ID {region_id} no encontrada'}, status=404)
    except Exception as e:
        logger.error(f"Error en obtener_datos_region: {str(e)}")
        return JsonResponse({'error': f'Error del servidor: {str(e)}'}, status=500)

@cache_page(60 * 240)  # 30 minutos  
def obtener_resumen_general(request):
    """API para obtener resumen general REAL de todas las regiones"""
    try:
        # Estadísticas generales REALES
        total_registros = DatosComercializacion.objects.count()
        total_regiones = Region.objects.count()
        total_productos = Producto.objects.count()
        total_mercados = Mercado.objects.count()

        # Fecha del registro más reciente
        ultimo_registro = DatosComercializacion.objects.order_by('-fecha').first()
        fecha_reciente = ultimo_registro.fecha if ultimo_registro else None

        # Estadísticas por región
        estadisticas_regiones = (DatosComercializacion.objects
                                .values('region__id_region', 'region__nombre')
                                .annotate(
                                    total_registros=Count('id'),
                                    total_productos=Count('producto', distinct=True),
                                    precio_promedio_global=Avg('precio_promedio')
                                )
                                .order_by('-total_registros'))

        datos = {
            'total_registros': total_registros,
            'total_regiones': total_regiones,
            'total_productos': total_productos,
            'total_mercados': total_mercados,
            'fecha_reciente': fecha_reciente.isoformat() if fecha_reciente else None,
            'estadisticas_regiones': list(estadisticas_regiones),
        }

        return JsonResponse(datos)

    except Exception as e:
        logger.error(f"Error en obtener_resumen_general: {str(e)}")
        return JsonResponse({'error': f'Error del servidor: {str(e)}'}, status=500)

@cache_page(60 * 240)  # 10 minutos  
def obtener_productos_region(request, region_id):
    try:
        with transaction.atomic():
            region = Region.objects.get(id_region=region_id)

            subsector = request.GET.get('subsector', '')
            producto = request.GET.get('producto', '')
            año = request.GET.get('año', '')

            queryset = DatosComercializacion.objects.filter(region=region)

            if subsector:
                queryset = queryset.filter(subsector__nombre__icontains=subsector)
            if producto:
                queryset = queryset.filter(producto__nombre__icontains=producto)
            if año:
                try:
                    queryset = queryset.filter(fecha__year=int(año))
                except ValueError:
                    pass

            #  CONTAR REGISTROS Y PRODUCTOS ÚNICOS CON FILTROS
            total_registros_filtrados = queryset.count()
            total_productos_filtrados = queryset.values('producto').distinct().count()
            total_mercados_filtrados = queryset.values('mercado').distinct().count()

            productos_data = (queryset
                             .values('producto__nombre', 'subsector__nombre')
                             .annotate(
                                 total_registros=Count('id'),
                                 precio_minimo=Min('precio_minimo'),
                                 precio_maximo=Max('precio_maximo'),
                                 precio_promedio=Avg('precio_promedio'),
                                 volumen_total=Sum('volumen')
                             )
                             .order_by('-volumen_total')[:50])

            datos = {
                'region': region.nombre,
                'filtros_aplicados': {
                    'subsector': subsector,
                    'producto': producto,
                    'año': año
                },
                'total_registros': total_registros_filtrados,    # ← Registros con filtros
                'total_productos': total_productos_filtrados,    # ← Productos únicos con filtros
                'total_mercados': total_mercados_filtrados,      # ← Mercados únicos con filtros
                'productos': list(productos_data)
            }

            return JsonResponse(datos)

    except Region.DoesNotExist:
        logger.warning(f"Region con ID {region_id} no encontrada")
        return JsonResponse({'error': f'Region con ID {region_id} no encontrada'}, status=404)
    except Exception as e:
        logger.error(f"Error en obtener_productos_region: {str(e)}")
        return JsonResponse({'error': f'Error del servidor: {str(e)}'}, status=500)


@cache_page(60 * 240)  # 10 minutos  
def obtener_filtros_disponibles(request):
    """API para obtener opciones de filtros disponibles"""
    try:
        with transaction.atomic():
            # Años disponibles
            años = (DatosComercializacion.objects
                    .dates('fecha', 'year', order='DESC')
                    .distinct())

            años_lista = [año.year for año in años]

            # Subsectores
            subsectores = (Subsector.objects
                          .annotate(total=Count('datoscomercializacion'))
                          .values('id', 'nombre')
                          .order_by('-total'))

            # Productos
            productos = (Producto.objects
                        .annotate(total=Count('datoscomercializacion'))
                        .values('id', 'nombre')
                        .order_by('nombre')
                        .distinct())

            datos = {
                'años': años_lista,
                'subsectores': list(subsectores),
                'productos': list(productos),
            }

            return JsonResponse(datos)

    except Exception as e:
        logger.error(f"Error en obtener_filtros_disponibles: {str(e)}")
        return JsonResponse({'error': f'Error del servidor: {str(e)}'}, status=500)

@cache_page(60 * 240)  # 10 minutos  
def obtener_productos_por_subsector(request, subsector_nombre):
    """API para obtener productos por nombre de subsector"""
    try:
        print(f"DEBUG: Buscando productos para subsector: {subsector_nombre}")

        productos = Producto.objects.filter(
            subsector__nombre__icontains=subsector_nombre
        ).values('id', 'nombre').order_by('nombre')

        print(f"DEBUG: Encontrados {productos.count()} productos")

        return JsonResponse({
            'productos': list(productos),
            'subsector_buscado': subsector_nombre,
            'total_productos': productos.count()
        })

    except Exception as e:
        print(f"ERROR en obtener_productos_por_subsector: {str(e)}")
        return JsonResponse({
            'error': f'Error del servidor: {str(e)}',
            'subsector_buscado': subsector_nombre
        }, status=500)


def debug_regiones(request):
    """Función temporal para debug de regiones"""
    regiones = Region.objects.all().values('id', 'id_region', 'nombre')
    datos_comerciales = DatosComercializacion.objects.values('region__id', 'region__id_region', 'region__nombre').annotate(
        total=Count('id')
    )

    return JsonResponse({
        'regiones': list(regiones),
        'datos_por_region': list(datos_comerciales)
    })
def comparar_regiones(request):
    """API para comparar multiples regiones"""
    try:
        region_ids = request.GET.get('regiones', '')
        if not region_ids:
            return JsonResponse({'error': 'No se especificaron regiones para comparar'}, status=400)

        region_ids = [int(id) for id in region_ids.split(',') if id.isdigit()]

        with transaction.atomic():
            datos_comparativos = []

            for region_id in region_ids:
                try:
                    region = Region.objects.get(id_region=region_id)

                    # Estadísticas básicas de la región
                    total_registros = DatosComercializacion.objects.filter(region=region).count()
                    productos_unicos = DatosComercializacion.objects.filter(region=region).values('producto').distinct().count()
                    mercados_unicos = DatosComercializacion.objects.filter(region=region).values('mercado').distinct().count()

                    # Estadísticas de precios y volumen
                    estadisticas = DatosComercializacion.objects.filter(region=region).aggregate(
                        precio_promedio=Avg('precio_promedio'),
                        volumen_total=Sum('volumen')
                    )

                    datos_comparativos.append({
                        'region_id': region.id_region,
                        'region_nombre': region.nombre,
                        'estadisticas': {
                            'total_registros': total_registros,
                            'productos_unicos': productos_unicos,
                            'total_mercados': mercados_unicos,
                            'precio_promedio': float(estadisticas['precio_promedio']) if estadisticas['precio_promedio'] else 0,
                            'volumen_total': float(estadisticas['volumen_total']) if estadisticas['volumen_total'] else 0
                        }
                    })

                except Region.DoesNotExist:
                    continue

            return JsonResponse({
                'regiones_comparadas': datos_comparativos,
                'total_regiones': len(datos_comparativos)
            })

    except Exception as e:
        logger.error(f"Error en comparar_regiones: {str(e)}")
        return JsonResponse({'error': f'Error del servidor: {str(e)}'}, status=500)

@cache_page(60 * 240)  # 30 minutos
def dashboard_analisis(request):
    """API principal para el dashboard - CON ANÁLISIS CONTEXTUAL"""
    try:
        # Obtener parámetros de filtro del request
        subsector = request.GET.get('subsector', '')
        producto = request.GET.get('producto', '')
        año = request.GET.get('año', '')
        region_id = request.GET.get('region_id', '')

        # CONVERTIR region_id a integer si existe
        if region_id and region_id.isdigit():
            region_id_int = int(region_id)
        else:
            region_id_int = None

        with transaction.atomic():
            # Determinar el tipo de análisis basado en los filtros
            tipo_analisis = determinar_tipo_analisis(region_id, producto, subsector)

            # Métricas principales contextuales
            metricas_principales = calcular_metricas_principales_filtradas(
                subsector=subsector,
                producto=producto,
                año=año,
                region_id=region_id_int  # ← Pasar el integer
            )

            # Análisis ESPECÍFICO según el contexto
            if tipo_analisis == 'region_especifica':
                # USAR region_id_int en lugar de region_id
                analisis_precios = calcular_analisis_precios_region(region_id_int, producto, subsector, año)
                analisis_volumenes = calcular_analisis_volumenes_region(region_id_int, producto, subsector, año)
                oportunidades_mercado = identificar_oportunidades_region(region_id_int, producto, subsector, año)

            elif tipo_analisis == 'producto_especifico':
                analisis_precios = calcular_analisis_precios_producto(producto, subsector, año)
                analisis_volumenes = calcular_analisis_volumenes_producto(producto, subsector, año)
                oportunidades_mercado = identificar_oportunidades_producto(producto, subsector, año)

            elif tipo_analisis == 'subsector_especifico':
                analisis_precios = calcular_analisis_precios_subsector(subsector, año)
                analisis_volumenes = calcular_analisis_volumenes_subsector(subsector, año)
                oportunidades_mercado = identificar_oportunidades_subsector(subsector, año)

            else:  # Vista nacional general
                analisis_precios = calcular_analisis_precios_filtrados(subsector, producto, año, region_id_int)  # ← integer
                analisis_volumenes = calcular_analisis_volumenes_filtrados(subsector, producto, año, region_id_int)  # ← integer
                oportunidades_mercado = identificar_oportunidades_mercado_filtradas(subsector, producto, año, region_id_int)  # ← integer

            # Análisis temporal siempre filtrado
            analisis_temporal = calcular_analisis_temporal_filtrado(subsector, producto, año, region_id_int)  # ← integer

            datos = {
                'metricas_principales': metricas_principales,
                'analisis_precios': analisis_precios,
                'analisis_volumenes': analisis_volumenes,
                'oportunidades_mercado': oportunidades_mercado,
                'analisis_temporal': analisis_temporal,
                'filtros_aplicados': {
                    'subsector': subsector,
                    'producto': producto,
                    'año': año,
                    'region_id': region_id
                },
                'tipo_analisis': tipo_analisis,
                'titulo_contextual': generar_titulo_contextual(tipo_analisis, region_id, producto, subsector)
            }

            return JsonResponse(datos)

    except Exception as e:
        logger.error(f"Error en dashboard_analisis: {str(e)}")
        return JsonResponse({
            'error': f'Error del servidor: {str(e)}',
            'metricas_principales': {},
            'analisis_precios': {'top_oportunidades': []},
            'analisis_volumenes': {'top_regiones_volumen': []},
            'oportunidades_mercado': [],
            'analisis_temporal': {'estacionalidad_productos': []},
            'filtros_aplicados': {},
            'tipo_analisis': 'general',
            'titulo_contextual': 'Análisis General'
        }, status=500)



    """API principal para el dashboard - CON ANÁLISIS CONTEXTUAL"""
    try:
        # Obtener parámetros de filtro del request
        subsector = request.GET.get('subsector', '')
        producto = request.GET.get('producto', '')
        año = request.GET.get('año', '')
        region_id = request.GET.get('region_id', '')

        with transaction.atomic():
            # Determinar el tipo de análisis basado en los filtros
            tipo_analisis = determinar_tipo_analisis(region_id, producto, subsector)

            # Métricas principales contextuales
            metricas_principales = calcular_metricas_principales_filtradas(
                subsector=subsector,
                producto=producto,
                año=año,
                region_id=region_id
            )

            # Análisis ESPECÍFICO según el contexto
            if tipo_analisis == 'region_especifica':
                analisis_precios = calcular_analisis_precios_region(region_id, producto, subsector, año)
                analisis_volumenes = calcular_analisis_volumenes_region(region_id, producto, subsector, año)
                oportunidades_mercado = identificar_oportunidades_region(region_id, producto, subsector, año)

            elif tipo_analisis == 'producto_especifico':
                analisis_precios = calcular_analisis_precios_producto(producto, subsector, año)
                analisis_volumenes = calcular_analisis_volumenes_producto(producto, subsector, año)
                oportunidades_mercado = identificar_oportunidades_producto(producto, subsector, año)

            elif tipo_analisis == 'subsector_especifico':
                analisis_precios = calcular_analisis_precios_subsector(subsector, año)
                analisis_volumenes = calcular_analisis_volumenes_subsector(subsector, año)
                oportunidades_mercado = identificar_oportunidades_subsector(subsector, año)

            else:  # Vista nacional general
                analisis_precios = calcular_analisis_precios_filtrados(subsector, producto, año, region_id)
                analisis_volumenes = calcular_analisis_volumenes_filtrados(subsector, producto, año, region_id)
                oportunidades_mercado = identificar_oportunidades_mercado_filtradas(subsector, producto, año, region_id)

            # Análisis temporal siempre filtrado
            analisis_temporal = calcular_analisis_temporal_filtrado(subsector, producto, año, region_id)

            datos = {
                'metricas_principales': metricas_principales,
                'analisis_precios': analisis_precios,
                'analisis_volumenes': analisis_volumenes,
                'oportunidades_mercado': oportunidades_mercado,
                'analisis_temporal': analisis_temporal,
                'filtros_aplicados': {
                    'subsector': subsector,
                    'producto': producto,
                    'año': año,
                    'region_id': region_id
                },
                'tipo_analisis': tipo_analisis,
                'titulo_contextual': generar_titulo_contextual(tipo_analisis, region_id, producto, subsector)
            }

            return JsonResponse(datos)

    except Exception as e:
        logger.error(f"Error en dashboard_analisis: {str(e)}")
        return JsonResponse({
            'error': f'Error del servidor: {str(e)}',
            'metricas_principales': {},
            'analisis_precios': {'top_oportunidades': []},
            'analisis_volumenes': {'top_regiones_volumen': []},
            'oportunidades_mercado': [],
            'analisis_temporal': {'estacionalidad_productos': []},
            'filtros_aplicados': {},
            'tipo_analisis': 'general',
            'titulo_contextual': 'Análisis General'
        }, status=500)


def calcular_analisis_precios_region(region_id, producto='', subsector='', año=''):
    """Análisis de precios ESPECÍFICO para una región"""
    try:
        print(f" DEBUG PRECIOS REGIONAL - region_id: {region_id}")

        if not region_id:
            return {'top_oportunidades': [], 'tipo_analisis': 'regional'}

        region = Region.objects.get(id_region=region_id)
        queryset = DatosComercializacion.objects.filter(region=region)

        #  FILTRAR OUTLIERS AGRESIVAMENTE
        queryset = queryset.filter(
            precio_promedio__gt=10,
            precio_promedio__lt=50000,
            precio_minimo__gt=5,
            precio_maximo__lt=100000,
            volumen__gt=0
        )

        print(f" DEBUG - Total registros región {region_id} (después de outliers): {queryset.count()}")

        if producto:
            queryset = queryset.filter(producto__nombre__icontains=producto)
        if subsector:
            queryset = queryset.filter(subsector__nombre__icontains=subsector)
        if año:
            try:
                queryset = queryset.filter(fecha__year=int(año))
            except ValueError:
                pass

        # Análisis de precios por mercado dentro de la región
        precios_por_mercado = (queryset
                              .values('mercado__nombre')
                              .annotate(
                                  precio_promedio=Avg('precio_promedio'),
                                  precio_min=Min('precio_promedio'),  # Usar promedio para consistencia
                                  precio_max=Max('precio_promedio'),  # Usar promedio para consistencia
                                  total_registros=Count('id')
                              )
                              .filter(
                                  total_registros__gt=0,
                                  precio_min__gt=10,
                                  precio_max__lt=50000
                              )
                              .order_by('-precio_promedio')[:10])

        print(f" DEBUG - Mercados con precios realistas: {len(precios_por_mercado)}")

        # Productos con mayor variación de precio en la región
        productos_variacion = (queryset
                              .values('producto__nombre')
                              .annotate(
                                  variacion_precio=((Max('precio_promedio') - Min('precio_promedio')) / Min('precio_promedio') * 100),
                                  precio_promedio=Avg('precio_promedio'),
                                  total_registros=Count('id')
                              )
                              .filter(
                                  total_registros__gt=1,
                                  variacion_precio__lt=300,  # Máximo 300% de variación
                                  precio_promedio__gt=10,
                                  precio_promedio__lt=50000
                              )
                              .order_by('-variacion_precio')[:8])

        print(f" DEBUG - Productos con variación realista: {len(productos_variacion)}")

        top_oportunidades = []

        # Oportunidades intra-regionales (entre mercados) - SOLO MÁRGENES RAZONABLES
        for mercado in precios_por_mercado[:5]:
            if (mercado['precio_max'] and mercado['precio_min'] and
                mercado['precio_min'] > 0 and mercado['precio_max'] > mercado['precio_min']):

                margen = ((mercado['precio_max'] - mercado['precio_min']) / mercado['precio_min'] * 100)

                # Solo incluir oportunidades con margen razonable (5% a 200%)
                if 5 <= margen <= 200:
                    top_oportunidades.append({
                        'producto': 'Múltiples productos',
                        'mejor_region': f"Mercado {mercado['mercado__nombre']}",
                        'peor_region': 'Otros mercados',
                        'precio_maximo': f"${mercado['precio_max']:.0f}",
                        'precio_minimo': f"${mercado['precio_min']:.0f}",
                        'diferencial_precio': f"${(mercado['precio_max'] - mercado['precio_min']):.0f}",
                        'margen_potencial': round(margen, 1)
                    })

        # Oportunidades por producto - SOLO MÁRGENES RAZONABLES
        for producto_data in productos_variacion[:3]:
            if 20 <= producto_data['variacion_precio'] <= 200:  # Entre 20% y 200%
                top_oportunidades.append({
                    'producto': producto_data['producto__nombre'],
                    'mejor_region': 'Mercados premium',
                    'peor_region': 'Mercados estándar',
                    'precio_maximo': 'Variable',
                    'precio_minimo': 'Variable',
                    'diferencial_precio': 'Oportunidad temporal',
                    'margen_potencial': round(producto_data['variacion_precio'], 1)
                })

        print(f" DEBUG - Oportunidades REALES encontradas: {len(top_oportunidades)}")

        return {
            'top_oportunidades': top_oportunidades,
            'tipo_analisis': 'regional',
            'descripcion': 'Análisis de oportunidades dentro de la región seleccionada'
        }

    except Region.DoesNotExist:
        print(f" DEBUG - Región {region_id} no existe")
        return {'top_oportunidades': [], 'tipo_analisis': 'regional'}
    except Exception as e:
        print(f" DEBUG - Error en análisis precios regional: {str(e)}")
        return {'top_oportunidades': [], 'tipo_analisis': 'regional'}




def calcular_analisis_volumenes_region(region_id, producto='', subsector='', año=''):
    """Análisis de volúmenes ESPECÍFICO para una región"""
    try:
        print(f" DEBUG VOLÚMENES REGIONAL - region_id: {region_id}")

        region = Region.objects.get(id_region=region_id)
        queryset = DatosComercializacion.objects.filter(region=region)

        if producto:
            queryset = queryset.filter(producto__nombre__icontains=producto)
        if subsector:
            queryset = queryset.filter(subsector__nombre__icontains=subsector)
        if año:
            try:
                queryset = queryset.filter(fecha__year=int(año))
            except ValueError:
                pass

        # Productos más transados en la región
        productos_volumen = (queryset
                           .values('producto__nombre', 'subsector__nombre')
                           .annotate(
                               volumen_total=Sum('volumen'),
                               precio_promedio=Avg('precio_promedio'),
                               total_registros=Count('id')
                           )
                           .order_by('-volumen_total')[:10])

        print(f" DEBUG - Productos con volumen: {len(productos_volumen)}")

        # Mercados más activos
        mercados_activos = (queryset
                           .values('mercado__nombre')
                           .annotate(
                               volumen_total=Sum('volumen'),
                               productos_unicos=Count('producto', distinct=True),
                               total_registros=Count('id')
                           )
                           .order_by('-volumen_total')[:5])

        print(f" DEBUG - Mercados activos: {len(mercados_activos)}")

        top_regiones_volumen = []

        # Convertir a formato esperado por el frontend
        for producto_data in productos_volumen[:5]:
            top_regiones_volumen.append({
                'region': 'Región actual',
                'producto_mas_transado': producto_data['producto__nombre'],
                'volumen_total': f"{producto_data['volumen_total']:,.0f}" if producto_data['volumen_total'] else '0',
                'mercados_activos': len(mercados_activos),
                'liquidez': round(producto_data['total_registros'] / max(len(mercados_activos), 1), 1)
            })

        print(f" DEBUG - Regiones volumen: {len(top_regiones_volumen)}")

        return {
            'top_regiones_volumen': top_regiones_volumen,
            'tipo_analisis': 'regional',
            'mercados_activos': list(mercados_activos)
        }

    except Region.DoesNotExist:
        print(f" DEBUG - Región {region_id} no existe")
        return {'top_regiones_volumen': [], 'tipo_analisis': 'regional'}
    except Exception as e:
        print(f" DEBUG - Error en análisis volúmenes regional: {str(e)}")
        return {'top_regiones_volumen': [], 'tipo_analisis': 'regional'}



def generar_titulo_contextual(tipo_analisis, region_id, producto, subsector):
    """Genera un título descriptivo basado en el contexto"""
    if tipo_analisis == 'region_especifica':
        try:
            region = Region.objects.get(id_region=int(region_id))
            return f"Análisis de {region.nombre}"
        except:
            return f"Análisis Regional"
    elif tipo_analisis == 'producto_especifico':
        return f"Análisis de {producto}"
    elif tipo_analisis == 'subsector_especifico':
        return f"Análisis del Subsector {subsector}"
    elif tipo_analisis == 'producto_region_especifica':
        try:
            region = Region.objects.get(id_region=int(region_id))
            return f"Análisis de {producto} en {region.nombre}"
        except:
            return f"Análisis Específico"
    else:
        return "Análisis Nacional General"


def determinar_tipo_analisis(region_id, producto, subsector):
    """Determina el tipo de análisis basado en los filtros aplicados"""
    if region_id and not producto and not subsector:
        return 'region_especifica'
    elif producto and not region_id:
        return 'producto_especifico'
    elif subsector and not region_id and not producto:
        return 'subsector_especifico'
    elif region_id and producto:
        return 'producto_region_especifica'
    else:
        return 'general'

def calcular_metricas_principales_filtradas(subsector='', producto='', año='', region_id=''):
    """Métricas calculadas con filtros aplicados"""
    try:
        print(f" DEBUG - Iniciando métricas con filtros: region_id={region_id}")

        # Base query con filtros
        queryset = DatosComercializacion.objects.all()

        if subsector:
            queryset = queryset.filter(subsector__nombre__icontains=subsector)
        if producto:
            queryset = queryset.filter(producto__nombre__icontains=producto)
        if año:
            try:
                queryset = queryset.filter(fecha__year=int(año))
            except ValueError:
                pass
        if region_id:
            try:
                region = Region.objects.get(id_region=int(region_id))
                queryset = queryset.filter(region=region)
            except (ValueError, Region.DoesNotExist):
                pass

        #  FILTRAR OUTLIERS AGRESIVAMENTE
        queryset = queryset.filter(
            precio_promedio__gt=10,        # Mínimo $10 (no regalados)
            precio_promedio__lt=50000,     # Máximo $50,000 (productos premium)
            precio_minimo__gt=5,           # Mínimo $5
            precio_minimo__lt=10000,       # Máximo $10,000
            precio_maximo__lt=100000,      # Máximo $100,000
            volumen__gt=0                  # Volumen positivo
        )

        total_registros = queryset.count()
        print(f" DEBUG - Total registros después de filtrar outliers: {total_registros}")

        if total_registros == 0:
            print(" DEBUG - No hay registros después de filtrar outliers")
            return {
                'margen_potencial_promedio': 0,
                'productos_alto_margen': 0,
                'regiones_activas': 0,
                'estacionalidad_promedio': 0,
                'total_registros_filtrados': 0
            }

        # Margen potencial promedio con filtros - USAR PRECIO_PROMEDIO
        productos_margen = (queryset
                           .values('producto__nombre')
                           .annotate(
                               precio_max=Max('precio_promedio'),
                               precio_min=Min('precio_promedio'),  # Usar promedio, no mínimo
                               total_reg=Count('id')
                           )
                           .filter(
                               precio_min__gt=10,      # Mínimo razonable
                               precio_max__lt=50000,   # Máximo razonable
                               total_reg__gt=1         # Al menos 2 registros
                           ))

        print(f" DEBUG - Productos para cálculo de margen: {productos_margen.count()}")

        margenes_validos = []
        for producto_data in productos_margen:
            try:
                margen = ((producto_data['precio_max'] - producto_data['precio_min']) /
                         producto_data['precio_min'] * 100)

                # FILTRAR MÁRGENES IRREALES
                if 5 < margen < 300:  # Margen entre 5% y 300% (más realista)
                    margenes_validos.append(margen)
                else:
                    print(f" DEBUG - Margen filtrado: {margen:.1f}% para {producto_data['producto__nombre']}")

            except (ZeroDivisionError, TypeError) as e:
                print(f" DEBUG - Error calculando margen: {e}")
                continue

        margen_promedio = sum(margenes_validos) / len(margenes_validos) if margenes_validos else 0

        # Productos con alto margen (>50%) con filtros
        productos_alto_margen = len([m for m in margenes_validos if m > 50])

        print(f" DEBUG - Margen promedio REAL: {margen_promedio:.1f}%")
        print(f" DEBUG - Margenes válidos: {len(margenes_validos)}")
        print(f" DEBUG - Productos con alto margen (>50%): {productos_alto_margen}")

        # Regiones activas con filtros
        regiones_activas = (queryset
                           .values('region__nombre')
                           .annotate(total_registros=Count('id'))
                           .filter(total_registros__gt=0)
                           .count())

        print(f" DEBUG - Regiones activas: {regiones_activas}")

        # Estacionalidad promedio con filtros
        try:
            variacion_mensual = (queryset
                                .values('fecha__month')
                                .annotate(
                                    avg_precio=Avg('precio_promedio'),
                                    total_reg=Count('id')
                                )
                                .filter(total_reg__gt=0))

            print(f" DEBUG - Meses con datos para estacionalidad: {len(variacion_mensual)}")

            if len(variacion_mensual) > 1:
                precios_mensuales = [item['avg_precio'] for item in variacion_mensual if item['avg_precio']]
                if precios_mensuales:
                    precio_max = max(precios_mensuales)
                    precio_min = min(precios_mensuales)
                    estacionalidad = ((precio_max - precio_min) / precio_min * 100) if precio_min > 0 else 0
                    print(f" DEBUG - Estacionalidad calculada: {estacionalidad:.1f}% (min: ${precio_min:.2f}, max: ${precio_max:.2f})")
                else:
                    estacionalidad = 0
                    print(" DEBUG - No hay precios mensuales válidos para estacionalidad")
            else:
                estacionalidad = 0
                print(" DEBUG - No hay suficientes meses para calcular estacionalidad")
        except Exception as e:
            estacionalidad = 0
            print(f" DEBUG - Error calculando estacionalidad: {e}")

        resultado = {
            'margen_potencial_promedio': round(margen_promedio, 1),
            'productos_alto_margen': productos_alto_margen,
            'regiones_activas': regiones_activas,
            'estacionalidad_promedio': round(estacionalidad, 1),
            'total_registros_filtrados': total_registros
        }

        print(f" DEBUG - Métricas finales REALES: {resultado}")
        print("=" * 50)

        return resultado

    except Exception as e:
        logger.error(f"Error en metricas principales filtradas: {str(e)}")
        print(f" DEBUG - ERROR GENERAL: {str(e)}")
        return {
            'margen_potencial_promedio': 0,
            'productos_alto_margen': 0,
            'regiones_activas': 0,
            'estacionalidad_promedio': 0,
            'total_registros_filtrados': 0
        }


def calcular_analisis_precios_filtrados(subsector='', producto='', año='', region_id=''):
    """Análisis de precios con filtros aplicados"""
    try:
        # Base query con filtros
        queryset = DatosComercializacion.objects.all()

        if subsector:
            queryset = queryset.filter(subsector__nombre__icontains=subsector)
        if producto:
            queryset = queryset.filter(producto__nombre__icontains=producto)
        if año:
            try:
                queryset = queryset.filter(fecha__year=int(año))
            except ValueError:
                pass
        if region_id:
            try:
                queryset = queryset.filter(region_id=int(region_id))
            except ValueError:
                pass

        # Productos con diferencial de precio - CON FILTROS
        productos_diferencial = (queryset
                               .values('producto__nombre')
                               .annotate(
                                   precio_maximo=Max('precio_promedio'),
                                   precio_minimo=Min('precio_promedio'),
                                   region_maximo=Max('region__nombre'),
                                   region_minimo=Min('region__nombre'),
                                   total_registros=Count('id'),
                                   avg_precio=Avg('precio_promedio')
                               )
                               .filter(
                                   total_registros__gt=1,
                                   precio_minimo__gt=0,
                                   precio_maximo__lt=1000000
                               )
                               .order_by('-precio_maximo')[:10])

        top_oportunidades = []
        for producto_data in productos_diferencial:
            try:
                margen_potencial = ((producto_data['precio_maximo'] - producto_data['precio_minimo']) /
                                   producto_data['precio_minimo'] * 100)

                if margen_potencial < 100000:
                    def formatear_precio(precio):
                        if precio >= 1000000:
                            return f"${precio/1000000:.1f}M"
                        elif precio >= 1000:
                            return f"${precio/1000:.1f}K"
                        return f"${precio:.0f}"

                    top_oportunidades.append({
                        'producto': producto_data['producto__nombre'],
                        'mejor_region': producto_data['region_maximo'],
                        'peor_region': producto_data['region_minimo'],
                        'precio_maximo': formatear_precio(producto_data['precio_maximo']),
                        'precio_minimo': formatear_precio(producto_data['precio_minimo']),
                        'diferencial_precio': formatear_precio(producto_data['precio_maximo'] - producto_data['precio_minimo']),
                        'margen_potencial': round(float(margen_potencial), 1)
                    })
            except (ZeroDivisionError, TypeError):
                continue

        return {
            'top_oportunidades': top_oportunidades
        }

    except Exception as e:
        logger.error(f"Error en análisis de precios filtrados: {str(e)}")
        return {'top_oportunidades': []}


def calcular_metricas_principales_reales():
    """Métricas calculadas solo con datos reales de la BD -  """
    try:
        #  Margen potencial promedio -
        productos_margen = (DatosComercializacion.objects
                           .values('producto__nombre')
                           .annotate(
                               precio_max=Max('precio_promedio'),
                               precio_min=Min('precio_promedio'),
                               total_reg=Count('id')
                           )
                           .filter(precio_min__gt=0, total_reg__gt=0))  #

        margenes_validos = []
        for producto in productos_margen:
            try:
                margen = ((producto['precio_max'] - producto['precio_min']) /
                         producto['precio_min'] * 100)
                if margen < 100000:  #   con outliers
                    margenes_validos.append(margen)
            except (ZeroDivisionError, TypeError):
                continue

        margen_promedio = sum(margenes_validos) / len(margenes_validos) if margenes_validos else 0

        #  Productos con alto margen (>20%) -
        productos_alto_margen = len([m for m in margenes_validos if m > 20])

        #  Regiones activas -
        regiones_activas = (DatosComercializacion.objects
                           .values('region__nombre')
                           .annotate(total_registros=Count('id'))
                           .filter(total_registros__gt=0)
                           .count())

        #  Estacionalidad promedio
        try:
            variacion_mensual = (DatosComercializacion.objects
                                .values('fecha__month')
                                .annotate(
                                    avg_precio=Avg('precio_promedio'),
                                    total_reg=Count('id')
                                )
                                .filter(total_reg__gt=0))

            if len(variacion_mensual) > 1:
                precios_mensuales = [item['avg_precio'] for item in variacion_mensual if item['avg_precio']]
                if precios_mensuales:
                    precio_max = max(precios_mensuales)
                    precio_min = min(precios_mensuales)
                    estacionalidad = ((precio_max - precio_min) / precio_min * 100) if precio_min > 0 else 0
                else:
                    estacionalidad = 0
            else:
                estacionalidad = 0
        except:
            estacionalidad = 0

        return {
            'margen_potencial_promedio': round(margen_promedio, 1),
            'productos_alto_margen': productos_alto_margen,
            'regiones_activas': regiones_activas,
            'estacionalidad_promedio': round(estacionalidad, 1),
        }
    except Exception as e:
        logger.error(f"Error en métricas principales: {str(e)}")
        return {}

def calcular_analisis_precios_reales():
    """Análisis de precios con datos reales - PARÁMETROS  S"""
    try:
        # Productos con diferencial de precio - PARÁMETROS  S
        productos_diferencial = (DatosComercializacion.objects
                               .values('producto__nombre')
                               .annotate(
                                   precio_maximo=Max('precio_promedio'),
                                   precio_minimo=Min('precio_promedio'),
                                   region_maximo=Max('region__nombre'),
                                   region_minimo=Min('region__nombre'),
                                   total_registros=Count('id'),
                                   avg_precio=Avg('precio_promedio')
                               )
                               .filter(
                                   total_registros__gt=1, #  : solo 1 registro
                                   precio_minimo__gt=0,
                                   precio_maximo__lt=1000000 #
                               )
                               .order_by('-precio_maximo')[:10]) # Más resultados

        top_oportunidades = []
        for producto in productos_diferencial:
            try:
                margen_potencial = ((producto['precio_maximo'] - producto['precio_minimo']) /
                                   producto['precio_minimo'] * 100)

                # Menos restrictivo con outliers
                if margen_potencial < 100000: #
                    # Formatear precios para mejor visualización
                    def formatear_precio(precio):
                        if precio >= 1000000:
                            return f"${precio/1000000:.1f}M"
                        elif precio >= 1000:
                            return f"${precio/1000:.1f}K"
                        return f"${precio:.0f}"

                    top_oportunidades.append({
                        'producto': producto['producto__nombre'],
                        'mejor_region': producto['region_maximo'],
                        'peor_region': producto['region_minimo'],
                        'precio_maximo': formatear_precio(producto['precio_maximo']),
                        'precio_minimo': formatear_precio(producto['precio_minimo']),
                        'diferencial_precio': formatear_precio(producto['precio_maximo'] - producto['precio_minimo']),
                        'margen_potencial': round(float(margen_potencial), 1)
                    })
            except (ZeroDivisionError, TypeError):
                continue

        return {
            'top_oportunidades': top_oportunidades
        }
    except Exception as e:
        logger.error(f"Error en análisis de precios: {str(e)}")
        return {'top_oportunidades': []}

def calcular_analisis_volumenes_reales():
    """Análisis de volúmenes con datos reales"""
    try:
        regiones_volumen = (DatosComercializacion.objects
                           .values('region__nombre')
                           .annotate(
                               volumen_total=Sum('volumen'),
                               productos_unicos=Count('producto', distinct=True),
                               mercados_activos=Count('mercado', distinct=True),
                               total_registros=Count('id')
                           )
                           .filter(volumen_total__isnull=False)
                           .order_by('-volumen_total')[:6])

        top_regiones_volumen = []
        for region in regiones_volumen:
            try:
                # Producto más transado en la región - DATO REAL
                producto_mas_transado = (DatosComercializacion.objects
                                       .filter(region__nombre=region['region__nombre'])
                                       .values('producto__nombre')
                                       .annotate(total_volumen=Sum('volumen'))
                                       .order_by('-total_volumen')
                                       .first())

                # Liquidez - CALCULO REAL
                liquidez = (region['total_registros'] / region['productos_unicos']
                           if region['productos_unicos'] > 0 else 0)

                # Formatear volumen
                def formatear_volumen(volumen):
                    if volumen >= 1000000:
                        return f"{volumen/1000000:.1f}M"
                    elif volumen >= 1000:
                        return f"{volumen/1000:.1f}K"
                    return f"{volumen:.0f}"

                top_regiones_volumen.append({
                    'region': region['region__nombre'],
                    'producto_mas_transado': (producto_mas_transado['producto__nombre']
                                             if producto_mas_transado else 'Sin datos'),
                    'volumen_total': formatear_volumen(float(region['volumen_total'] or 0)),
                    'mercados_activos': region['mercados_activos'],
                    'liquidez': round(liquidez, 1)
                })
            except (TypeError, KeyError):
                continue

        return {
            'top_regiones_volumen': top_regiones_volumen
        }
    except Exception as e:
        logger.error(f"Error en análisis de volúmenes: {str(e)}")
        return {'top_regiones_volumen': []}

def identificar_oportunidades_mercado_reales():
    """Oportunidades basadas en datos reales -  """
    try:
        oportunidades = []

        #  Productos con variación de precios -
        productos_variacion = (DatosComercializacion.objects
                              .values('producto__nombre')
                              .annotate(
                                  variacion=((Max('precio_promedio') - Min('precio_promedio')) / Min('precio_promedio') * 100),
                                  total_registros=Count('id'),
                                  avg_precio=Avg('precio_promedio')
                              )
                              .filter(
                                  variacion__gt=10, #  : 10% en lugar de 20%
                                  variacion__lt=10000, #
                                  total_registros__gt=1, #
                                  avg_precio__lt=50000 #
                              )
                              .order_by('-variacion')[:5]) # Más resultados

        for producto in productos_variacion:
            potencial = 'Alto' if producto['variacion'] > 100 else 'Medio' if producto['variacion'] > 30 else 'Bajo'

            oportunidades.append({
                'tipo': 'Arbitraje Regional',
                'descripcion': f"{producto['producto__nombre']} - Variación {producto['variacion']:.1f}%",
                'potencial': potencial,
                'riesgo': 'Medio',
                'recomendacion': 'Comprar en regiones de bajo precio, vender en regiones de alto precio'
            })

        #  Productos con buen volumen y precio estable
        productos_volumen = (DatosComercializacion.objects
                           .values('producto__nombre')
                           .annotate(
                               volumen_total=Sum('volumen'),
                               precio_promedio=Avg('precio_promedio'),
                               total_registros=Count('id'),
                               regiones_activas=Count('region', distinct=True)
                           )
                           .filter(
                               volumen_total__gt=1000,
                               total_registros__gt=2,
                               regiones_activas__gt=1
                           )
                           .order_by('-volumen_total')[:3])

        for producto in productos_volumen:
            oportunidades.append({
                'tipo': 'Mercado Estable',
                'descripcion': f"{producto['producto__nombre']} - Alto volumen ({producto['volumen_total']:.0f})",
                'potencial': 'Medio',
                'riesgo': 'Bajo',
                'recomendacion': 'Mercado estable con buena liquidez'
            })

        return oportunidades

    except Exception as e:
        logger.error(f"Error identificando oportunidades: {str(e)}")
        return []
def calcular_analisis_volumenes_filtrados(subsector='', producto='', año='', region_id=''):
    """Análisis de volúmenes con filtros aplicados"""
    try:
        # Base query con filtros
        queryset = DatosComercializacion.objects.all()

        if subsector:
            queryset = queryset.filter(subsector__nombre__icontains=subsector)
        if producto:
            queryset = queryset.filter(producto__nombre__icontains=producto)
        if año:
            try:
                queryset = queryset.filter(fecha__year=int(año))
            except ValueError:
                pass
        if region_id:
            try:
                queryset = queryset.filter(region_id=int(region_id))
            except ValueError:
                pass

        # Regiones con volumen - CON FILTROS
        regiones_volumen = (queryset
                           .values('region__nombre')
                           .annotate(
                               volumen_total=Sum('volumen'),
                               productos_unicos=Count('producto', distinct=True),
                               mercados_activos=Count('mercado', distinct=True),
                               total_registros=Count('id')
                           )
                           .filter(volumen_total__isnull=False)
                           .order_by('-volumen_total')[:6])

        top_regiones_volumen = []
        for region in regiones_volumen:
            try:
                # Producto más transado en la región - CON FILTROS
                producto_mas_transado_query = queryset.filter(region__nombre=region['region__nombre'])
                if producto:
                    producto_mas_transado_query = producto_mas_transado_query.filter(producto__nombre__icontains=producto)

                producto_mas_transado = (producto_mas_transado_query
                                       .values('producto__nombre')
                                       .annotate(total_volumen=Sum('volumen'))
                                       .order_by('-total_volumen')
                                       .first())

                # Liquidez - CALCULO CON FILTROS
                liquidez = (region['total_registros'] / region['productos_unicos']
                           if region['productos_unicos'] > 0 else 0)

                # Formatear volumen
                def formatear_volumen(volumen):
                    if volumen >= 1000000:
                        return f"{volumen/1000000:.1f}M"
                    elif volumen >= 1000:
                        return f"{volumen/1000:.1f}K"
                    return f"{volumen:.0f}"

                top_regiones_volumen.append({
                    'region': region['region__nombre'],
                    'producto_mas_transado': (producto_mas_transado['producto__nombre']
                                             if producto_mas_transado else 'Sin datos'),
                    'volumen_total': formatear_volumen(float(region['volumen_total'] or 0)),
                    'mercados_activos': region['mercados_activos'],
                    'liquidez': round(liquidez, 1)
                })
            except (TypeError, KeyError):
                continue

        return {
            'top_regiones_volumen': top_regiones_volumen
        }

    except Exception as e:
        logger.error(f"Error en análisis de volúmenes filtrados: {str(e)}")
        return {'top_regiones_volumen': []}

def identificar_oportunidades_mercado_filtradas(subsector='', producto='', año='', region_id=''):
    """Oportunidades basadas en datos filtrados"""
    try:
        # Base query con filtros
        queryset = DatosComercializacion.objects.all()

        if subsector:
            queryset = queryset.filter(subsector__nombre__icontains=subsector)
        if producto:
            queryset = queryset.filter(producto__nombre__icontains=producto)
        if año:
            try:
                queryset = queryset.filter(fecha__year=int(año))
            except ValueError:
                pass
        if region_id:
            try:
                queryset = queryset.filter(region_id=int(region_id))
            except ValueError:
                pass

        oportunidades = []

        # Productos con variación de precios - CON FILTROS
        productos_variacion = (queryset
                              .values('producto__nombre')
                              .annotate(
                                  variacion=((Max('precio_promedio') - Min('precio_promedio')) / Min('precio_promedio') * 100),
                                  total_registros=Count('id'),
                                  avg_precio=Avg('precio_promedio')
                              )
                              .filter(
                                  variacion__gt=10,
                                  variacion__lt=10000,
                                  total_registros__gt=1,
                                  avg_precio__lt=50000
                              )
                              .order_by('-variacion')[:5])

        for producto_data in productos_variacion:
            potencial = 'Alto' if producto_data['variacion'] > 100 else 'Medio' if producto_data['variacion'] > 30 else 'Bajo'

            oportunidades.append({
                'tipo': 'Arbitraje Regional',
                'descripcion': f"{producto_data['producto__nombre']} - Variación {producto_data['variacion']:.1f}%",
                'potencial': potencial,
                'riesgo': 'Medio',
                'recomendacion': 'Comprar en regiones de bajo precio, vender en regiones de alto precio'
            })

        # Productos con buen volumen y precio estable - CON FILTROS
        productos_volumen = (queryset
                           .values('producto__nombre')
                           .annotate(
                               volumen_total=Sum('volumen'),
                               precio_promedio=Avg('precio_promedio'),
                               total_registros=Count('id'),
                               regiones_activas=Count('region', distinct=True)
                           )
                           .filter(
                               volumen_total__gt=1000,
                               total_registros__gt=2,
                               regiones_activas__gt=1
                           )
                           .order_by('-volumen_total')[:3])

        for producto_data in productos_volumen:
            oportunidades.append({
                'tipo': 'Mercado Estable',
                'descripcion': f"{producto_data['producto__nombre']} - Alto volumen ({producto_data['volumen_total']:.0f})",
                'potencial': 'Medio',
                'riesgo': 'Bajo',
                'recomendacion': 'Mercado estable con buena liquidez'
            })

        return oportunidades

    except Exception as e:
        logger.error(f"Error identificando oportunidades filtradas: {str(e)}")
        return []

def calcular_analisis_temporal_filtrado(subsector='', producto='', año='', region_id=''):
    """Análisis temporal con filtros aplicados"""
    try:
        print(f" DEBUG TEMPORAL - Iniciando con region_id: {region_id}")

        # Base query con filtros
        queryset = DatosComercializacion.objects.all()

        #  FILTRAR OUTLIERS
        queryset = queryset.filter(
            precio_promedio__gt=10,
            precio_promedio__lt=50000,
            volumen__gt=0
        )

        if subsector:
            queryset = queryset.filter(subsector__nombre__icontains=subsector)
        if producto:
            queryset = queryset.filter(producto__nombre__icontains=producto)
        if año:
            try:
                queryset = queryset.filter(fecha__year=int(año))
            except ValueError:
                pass
        if region_id:
            try:
                region = Region.objects.get(id_region=int(region_id))
                queryset = queryset.filter(region=region)
                print(f" DEBUG TEMPORAL - Filtrado por región: {region.nombre}")
            except (ValueError, Region.DoesNotExist):
                print(f" DEBUG TEMPORAL - Región {region_id} no encontrada")
                pass

        total_registros = queryset.count()
        print(f" DEBUG TEMPORAL - Total registros después de filtros: {total_registros}")

        if total_registros == 0:
            print(" DEBUG TEMPORAL - No hay datos después de filtrar")
            return {'estacionalidad_productos': []}

        #  RELAJAR FILTROS para estacionalidad
        productos_estacionalidad = (queryset
                                  .values('producto__nombre')
                                  .annotate(
                                      total_registros=Count('id'),
                                      variacion_mensual=((Max('precio_promedio') - Min('precio_promedio')) / Min('precio_promedio') * 100),
                                      meses_activos=Count('fecha__month', distinct=True)
                                  )
                                  .filter(
                                      total_registros__gt=1,        #  BAJAR de 2 a 1
                                      variacion_mensual__lt=1000,   # Mantener filtro de outliers
                                      meses_activos__gt=0           #  BAJAR de 1 a 0
                                  )
                                  .order_by('-variacion_mensual')[:10])  #  AUMENTAR resultados

        print(f" DEBUG TEMPORAL - Productos para estacionalidad: {len(productos_estacionalidad)}")

        estacionalidad_productos = []

        for producto_data in productos_estacionalidad:
            # Análisis básico por mes - CON FILTROS
            meses_query = queryset.filter(producto__nombre=producto_data['producto__nombre'])
            meses_data = (meses_query
                         .values('fecha__month')
                         .annotate(
                             avg_precio=Avg('precio_promedio'),
                             total_reg=Count('id')
                         )
                         .order_by('fecha__month'))

            print(f" DEBUG TEMPORAL - Producto {producto_data['producto__nombre']}: {len(meses_data)} meses de datos")

            if len(meses_data) >= 2:  #  Necesitamos al menos 2 meses para análisis
                # Encontrar mes con precio más bajo y más alto
                try:
                    mes_min = min(meses_data, key=lambda x: x['avg_precio'] or 0)
                    mes_max = max(meses_data, key=lambda x: x['avg_precio'] or 0)

                    meses_nombres = {
                        1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
                        5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
                        9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
                    }

                    mejor_comprar = meses_nombres.get(mes_min['fecha__month'], 'Analizar')
                    mejor_vender = meses_nombres.get(mes_max['fecha__month'], 'Analizar')
                except (ValueError, TypeError):
                    mejor_comprar = 'Analizar'
                    mejor_vender = 'Analizar'
            else:
                mejor_comprar = 'Analizar'
                mejor_vender = 'Analizar'
                print(f" DEBUG TEMPORAL - {producto_data['producto__nombre']}: No hay suficientes meses")

            # Determinar tendencia - CON FILTROS
            tendencia = 'Estable'  # Por defecto
            try:
                tendencia_data = (meses_query
                                .order_by('fecha'))

                if tendencia_data.count() >= 2:
                    primer_precio = tendencia_data.first().precio_promedio or 0
                    ultimo_precio = tendencia_data.last().precio_promedio or 0

                    if primer_precio > 0:
                        tendencia_porcentaje = ((ultimo_precio - primer_precio) / primer_precio * 100)
                        if tendencia_porcentaje > 10:  #  Aumentar sensibilidad
                            tendencia = 'Alcista'
                        elif tendencia_porcentaje < -10:
                            tendencia = 'Bajista'
                        else:
                            tendencia = 'Estable'
            except Exception as e:
                print(f" DEBUG TEMPORAL - Error en tendencia: {e}")
                tendencia = 'Estable'

            estacionalidad_productos.append({
                'producto': producto_data['producto__nombre'],
                'mejor_mes_comprar': mejor_comprar,
                'mejor_mes_vender': mejor_vender,
                'variacion_estacional': round(producto_data['variacion_mensual'], 1),
                'tendencia': tendencia
            })

        print(f" DEBUG TEMPORAL - Estacionalidad productos final: {len(estacionalidad_productos)}")

        return {
            'estacionalidad_productos': estacionalidad_productos
        }

    except Exception as e:
        print(f" DEBUG TEMPORAL - ERROR: {str(e)}")
        return {'estacionalidad_productos': []}



def calcular_analisis_temporal_reales():
    """Análisis temporal con datos reales -  """
    try:
        # Calcular estacionalidad real por producto -
        productos_estacionalidad = (DatosComercializacion.objects
                                  .values('producto__nombre')
                                  .annotate(
                                      total_registros=Count('id'),
                                      variacion_mensual=((Max('precio_promedio') - Min('precio_promedio')) / Min('precio_promedio') * 100),
                                      meses_activos=Count('fecha__month', distinct=True)
                                  )
                                  .filter(
                                      total_registros__gt=2, #
                                      variacion_mensual__lt=100000, #
                                      meses_activos__gt=1 # Al menos 2 meses diferentes
                                  )
                                  .order_by('-variacion_mensual')[:8]) # Más resultados

        estacionalidad_productos = []
        for producto in productos_estacionalidad:
            # Análisis básico por mes
            meses_data = (DatosComercializacion.objects
                         .filter(producto__nombre=producto['producto__nombre'])
                         .values('fecha__month')
                         .annotate(
                             avg_precio=Avg('precio_promedio'),
                             total_reg=Count('id')
                         )
                         .order_by('fecha__month'))

            if meses_data:
                # Encontrar mes con precio más bajo y más alto
                mes_min = min(meses_data, key=lambda x: x['avg_precio'])
                mes_max = max(meses_data, key=lambda x: x['avg_precio'])

                meses_nombres = {
                    1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
                    5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
                    9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
                }

                mejor_comprar = meses_nombres.get(mes_min['fecha__month'], 'Analizar')
                mejor_vender = meses_nombres.get(mes_max['fecha__month'], 'Analizar')
            else:
                mejor_comprar = 'Analizar'
                mejor_vender = 'Analizar'

            # Determinar tendencia
            tendencia_data = (DatosComercializacion.objects
                            .filter(producto__nombre=producto['producto__nombre'])
                            .order_by('fecha'))

            if tendencia_data.count() > 1:
                primer_precio = tendencia_data.first().precio_promedio
                ultimo_precio = tendencia_data.last().precio_promedio

                if primer_precio and ultimo_precio and primer_precio > 0:
                    tendencia_porcentaje = ((ultimo_precio - primer_precio) / primer_precio * 100)
                    if tendencia_porcentaje > 5:
                        tendencia = 'Alcista'
                    elif tendencia_porcentaje < -5:
                        tendencia = 'Bajista'
                    else:
                        tendencia = 'Estable'
                else:
                    tendencia = 'Estable'
            else:
                tendencia = 'Estable'

            estacionalidad_productos.append({
                'producto': producto['producto__nombre'],
                'mejor_mes_comprar': mejor_comprar,
                'mejor_mes_vender': mejor_vender,
                'variacion_estacional': round(producto['variacion_mensual'], 1),
                'tendencia': tendencia
            })

        return {
            'estacionalidad_productos': estacionalidad_productos
        }
    except Exception as e:
        logger.error(f"Error en análisis temporal: {str(e)}")
        return {'estacionalidad_productos': []}

def estadisticas_regiones(request):
    """API para obtener estadisticas comparativas de todas las regiones"""
    try:
        with transaction.atomic():
            estadisticas = (DatosComercializacion.objects
                           .values('region__id_region', 'region__nombre')
                           .annotate(
                               total_registros=Count('id'),
                               total_productos=Count('producto', distinct=True),
                               total_mercados=Count('mercado', distinct=True),
                               precio_promedio=Avg('precio_promedio'),
                               volumen_total=Sum('volumen')
                           )
                           .order_by('region__id_region'))

            return JsonResponse({
                'estadisticas': list(estadisticas),
                'total_regiones': len(estadisticas)
            })

    except Exception as e:
        logger.error(f"Error en estadisticas_regiones: {str(e)}")
        return JsonResponse({'error': f'Error del servidor: {str(e)}'}, status=500)

def estadisticas_productos(request):
    """API para obtener estadisticas de productos a nivel nacional"""
    try:
        with transaction.atomic():
            producto = request.GET.get('producto', '')
            subsector = request.GET.get('subsector', '')
            año = request.GET.get('año', '')

            queryset = DatosComercializacion.objects.all()

            if producto:
                queryset = queryset.filter(producto__nombre__icontains=producto)
            if subsector:
                queryset = queryset.filter(subsector__nombre__icontains=subsector)
            if año:
                try:
                    queryset = queryset.filter(fecha__year=int(año))
                except ValueError:
                    pass

            estadisticas = (queryset
                           .values('producto__nombre', 'subsector__nombre')
                           .annotate(
                               total_registros=Count('id'),
                               regiones_activas=Count('region', distinct=True),
                               precio_minimo=Min('precio_minimo'),
                               precio_maximo=Max('precio_maximo'),
                               precio_promedio=Avg('precio_promedio'),
                               volumen_total=Sum('volumen')
                           )
                           .order_by('-volumen_total')[:100])

            return JsonResponse({
                'estadisticas': list(estadisticas),
                'filtros_aplicados': {
                    'producto': producto,
                    'subsector': subsector,
                    'año': año
                }
            })

    except Exception as e:
        logger.error(f"Error en estadisticas_productos: {str(e)}")
        return JsonResponse({'error': f'Error del servidor: {str(e)}'}, status=500)

def exportar_dashboard_pdf(request):
    """Exportar datos del dashboard a PDF"""
    try:
        # Obtener datos reales del dashboard
        datos = obtener_datos_dashboard_reales()

        if not datos:
            return HttpResponse("No hay datos disponibles para exportar", status=404)

        # Crear buffer para el PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch)
        elements = []
        styles = getSampleStyleSheet()

        # Título del documento
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1,
            textColor=colors.HexColor('#2c5aa0')
        )

        title = Paragraph("Reporte de Análisis - AgroPrecios Chile", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.2*inch))

        #  Métricas Principales
        elements.append(Paragraph("Métricas Principales", styles['Heading2']))
        if datos.get('metricas_principales'):
            metricas_data = [
                ['Métrica', 'Valor'],
                ['Margen Potencial Promedio', f"{datos['metricas_principales'].get('margen_potencial_promedio', 0)}%"],
                ['Productos con Alto Margen', datos['metricas_principales'].get('productos_alto_margen', 0)],
                ['Regiones Activas', datos['metricas_principales'].get('regiones_activas', 0)],
                ['Estacionalidad Promedio', f"{datos['metricas_principales'].get('estacionalidad_promedio', 0)}%"]
            ]
            metricas_table = Table(metricas_data, colWidths=[3*inch, 2*inch])
            metricas_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c5aa0')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8f9fa')),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ]))
            elements.append(metricas_table)
            elements.append(Spacer(1, 0.3*inch))

        #  Oportunidades de Precios
        if datos.get('analisis_precios') and datos['analisis_precios'].get('top_oportunidades'):
            elements.append(Paragraph("Oportunidades de Precios por Región", styles['Heading2']))
            oportunidades_data = [['Producto', 'Mejor Región', 'Precio Max', 'Peor Región', 'Precio Min', 'Margen %']]

            for op in datos['analisis_precios']['top_oportunidades'][:5]:
                oportunidades_data.append([
                    op.get('producto', ''),
                    op.get('mejor_region', ''),
                    op.get('precio_maximo', ''),
                    op.get('peor_region', ''),
                    op.get('precio_minimo', ''),
                    f"{op.get('margen_potencial', 0)}%"
                ])

            op_table = Table(oportunidades_data, colWidths=[1.2*inch, 1.2*inch, 1*inch, 1.2*inch, 1*inch, 0.8*inch])
            op_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#28a745')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ]))
            elements.append(op_table)
            elements.append(Spacer(1, 0.3*inch))

        #  Análisis de Volúmenes
        if datos.get('analisis_volumenes') and datos['analisis_volumenes'].get('top_regiones_volumen'):
            elements.append(Paragraph("Análisis de Volúmenes Comerciales", styles['Heading2']))
            volumenes_data = [['Región', 'Producto Principal', 'Volumen Total', 'Mercados', 'Liquidez']]

            for vol in datos['analisis_volumenes']['top_regiones_volumen'][:5]:
                volumenes_data.append([
                    vol.get('region', ''),
                    vol.get('producto_mas_transado', ''),
                    vol.get('volumen_total', ''),
                    vol.get('mercados_activos', ''),
                    vol.get('liquidez', '')
                ])

            vol_table = Table(volumenes_data, colWidths=[1.5*inch, 1.5*inch, 1.2*inch, 0.8*inch, 0.8*inch])
            vol_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#17a2b8')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ]))
            elements.append(vol_table)

        # Construir el PDF
        doc.build(elements)

        # Preparar respuesta
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="reporte_analisis_agroprecios.pdf"'

        return response

    except Exception as e:
        logger.error(f"Error generando PDF: {str(e)}")
        return HttpResponse(f"Error al generar el PDF: {str(e)}", status=500)


def exportar_dashboard_excel(request):
    """Exportar datos del dashboard a Excel"""
    try:
        # Obtener datos reales del dashboard
        datos = obtener_datos_dashboard_reales()

        if not datos:
            return HttpResponse("No hay datos disponibles para exportar", status=404)

        # Crear workbook
        wb = openpyxl.Workbook()

        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=14, color="2c5aa0")
        subheader_font = Font(bold=True, size=11, color="2c5aa0")
        normal_font = Font(size=10)

        header_fill_blue = PatternFill(start_color="2c5aa0", end_color="2c5aa0", fill_type="solid")
        header_fill_green = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
        header_fill_teal = PatternFill(start_color="17a2b8", end_color="17a2b8", fill_type="solid")
        header_fill_orange = PatternFill(start_color="fd7e14", end_color="fd7e14", fill_type="solid")

        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')

        #  HOJA RESUMEN - Métricas Principales
        ws1 = wb.active
        ws1.title = "Resumen Ejecutivo"

        # Título principal
        ws1['A1'] = "REPORTE DE ANÁLISIS - AGROPRECIOS CHILE"
        ws1['A1'].font = title_font
        ws1.merge_cells('A1:F1')
        ws1.row_dimensions[1].height = 25

        # Fecha de generación
        from datetime import datetime
        ws1['A2'] = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws1['A2'].font = Font(italic=True, size=9, color="666666")
        ws1.merge_cells('A2:F2')

        # Métricas Principales
        ws1['A4'] = "MÉTRICAS PRINCIPALES"
        ws1['A4'].font = subheader_font
        ws1.merge_cells('A4:F4')

        if datos.get('metricas_principales'):
            metricas = datos['metricas_principales']
            headers = ['Métrica', 'Valor', 'Interpretación']
            metricas_data = [
                ['Margen Potencial Promedio', f"{metricas.get('margen_potencial_promedio', 0)}%", 'Oportunidad de arbitraje promedio entre regiones'],
                ['Productos con Alto Margen', metricas.get('productos_alto_margen', 0), 'Productos con variación de precio >30%'],
                ['Regiones Activas', metricas.get('regiones_activas', 0), 'Regiones con datos comerciales'],
                ['Estacionalidad Promedio', f"{metricas.get('estacionalidad_promedio', 0)}%", 'Variación promedio estacional']
            ]

            # Escribir headers
            for col, header in enumerate(headers, 1):
                cell = ws1.cell(row=6, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill_blue
                cell.alignment = center_align

            # Escribir datos
            for row_idx, fila in enumerate(metricas_data, 7):
                for col_idx, valor in enumerate(fila, 1):
                    cell = ws1.cell(row=row_idx, column=col_idx, value=valor)
                    cell.font = normal_font
                    cell.alignment = left_align if col_idx == 3 else center_align

        #  HOJA OPORTUNIDADES DE PRECIOS
        ws2 = wb.create_sheet("Oportunidades de Precios")

        ws2['A1'] = "OPORTUNIDADES DE PRECIOS POR REGIÓN"
        ws2['A1'].font = title_font
        ws2.merge_cells('A1:H1')

        if datos.get('analisis_precios') and datos['analisis_precios'].get('top_oportunidades'):
            oportunidades = datos['analisis_precios']['top_oportunidades']

            headers = ['Producto', 'Mejor Región', 'Precio Máximo', 'Peor Región', 'Precio Mínimo', 'Diferencial', 'Margen %', 'Recomendación']

            # Escribir headers
            for col, header in enumerate(headers, 1):
                cell = ws2.cell(row=3, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill_green
                cell.alignment = center_align

            # Escribir datos
            for row_idx, op in enumerate(oportunidades, 4):
                recomendacion = f"Comprar en {op.get('peor_region', '')} - Vender en {op.get('mejor_region', '')}"

                ws2.cell(row=row_idx, column=1, value=op.get('producto', '')).font = normal_font
                ws2.cell(row=row_idx, column=2, value=op.get('mejor_region', '')).font = normal_font
                ws2.cell(row=row_idx, column=3, value=op.get('precio_maximo', '')).font = normal_font
                ws2.cell(row=row_idx, column=4, value=op.get('peor_region', '')).font = normal_font
                ws2.cell(row=row_idx, column=5, value=op.get('precio_minimo', '')).font = normal_font
                ws2.cell(row=row_idx, column=6, value=op.get('diferencial_precio', '')).font = normal_font
                ws2.cell(row=row_idx, column=7, value=op.get('margen_potencial', 0)).font = normal_font
                ws2.cell(row=row_idx, column=8, value=recomendacion).font = Font(size=9, color="2c5aa0")

        #  HOJA ANÁLISIS DE VOLÚMENES
        ws3 = wb.create_sheet("Análisis de Volúmenes")

        ws3['A1'] = "ANÁLISIS DE VOLÚMENES COMERCIALES"
        ws3['A1'].font = title_font
        ws3.merge_cells('A1:G1')

        if datos.get('analisis_volumenes') and datos['analisis_volumenes'].get('top_regiones_volumen'):
            volumenes = datos['analisis_volumenes']['top_regiones_volumen']

            headers = ['Región', 'Producto Principal', 'Volumen Total', 'Mercados Activos', 'Liquidez', 'Potencial', 'Observaciones']

            # Escribir headers
            for col, header in enumerate(headers, 1):
                cell = ws3.cell(row=3, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill_teal
                cell.alignment = center_align

            # Escribir datos
            for row_idx, vol in enumerate(volumenes, 4):
                liquidez = float(vol.get('liquidez', 0))
                potencial = "Alto" if liquidez > 10 else "Medio" if liquidez > 5 else "Bajo"
                observaciones = "Mercado muy líquido" if liquidez > 10 else "Mercado estable" if liquidez > 5 else "Mercado emergente"

                ws3.cell(row=row_idx, column=1, value=vol.get('region', '')).font = normal_font
                ws3.cell(row=row_idx, column=2, value=vol.get('producto_mas_transado', '')).font = normal_font
                ws3.cell(row=row_idx, column=3, value=vol.get('volumen_total', '')).font = normal_font
                ws3.cell(row=row_idx, column=4, value=vol.get('mercados_activos', '')).font = normal_font
                ws3.cell(row=row_idx, column=5, value=vol.get('liquidez', '')).font = normal_font
                ws3.cell(row=row_idx, column=6, value=potencial).font = normal_font
                ws3.cell(row=row_idx, column=7, value=observaciones).font = Font(size=9, color="666666")

        #  HOJA OPORTUNIDADES DE MERCADO
        ws4 = wb.create_sheet("Oportunidades de Mercado")

        ws4['A1'] = "OPORTUNIDADES ESTRATÉGICAS DE MERCADO"
        ws4['A1'].font = title_font
        ws4.merge_cells('A1:E1')

        if datos.get('oportunidades_mercado'):
            oportunidades = datos['oportunidades_mercado']

            headers = ['Tipo Oportunidad', 'Descripción', 'Potencial', 'Riesgo', 'Recomendación']

            # Escribir headers
            for col, header in enumerate(headers, 1):
                cell = ws4.cell(row=3, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill_orange
                cell.alignment = center_align

            # Escribir datos
            for row_idx, op in enumerate(oportunidades, 4):
                ws4.cell(row=row_idx, column=1, value=op.get('tipo', '')).font = normal_font
                ws4.cell(row=row_idx, column=2, value=op.get('descripcion', '')).font = normal_font
                ws4.cell(row=row_idx, column=3, value=op.get('potencial', '')).font = normal_font
                ws4.cell(row=row_idx, column=4, value=op.get('riesgo', '')).font = normal_font
                ws4.cell(row=row_idx, column=5, value=op.get('recomendacion', '')).font = Font(size=9, color="2c5aa0")

        # AJUSTAR ANCHO DE COLUMNAS EN TODAS LAS HOJAS
        for ws in [ws1, ws2, ws3, ws4]:
            if ws == ws1:
                ws.column_dimensions['A'].width = 25
                ws.column_dimensions['B'].width = 15
                ws.column_dimensions['C'].width = 35
            elif ws == ws2:
                for col in range(1, 9):
                    ws.column_dimensions[get_column_letter(col)].width = 18
            elif ws == ws3:
                for col in range(1, 8):
                    ws.column_dimensions[get_column_letter(col)].width = 16
            elif ws == ws4:
                for col in range(1, 6):
                    ws.column_dimensions[get_column_letter(col)].width = 20

        # Aplicar bordes a todas las celdas con datos
        for ws in [ws1, ws2, ws3, ws4]:
            thin_border = Border(left=Side(style='thin'),
                               right=Side(style='thin'),
                               top=Side(style='thin'),
                               bottom=Side(style='thin'))

            for row in ws.iter_rows():
                for cell in row:
                    if cell.value:
                        cell.border = thin_border

        # Preparar respuesta
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="reporte_completo_agroprecios.xlsx"'

        return response

    except Exception as e:
        logger.error(f"Error generando Excel: {str(e)}")
        return HttpResponse(f"Error al generar el Excel: {str(e)}", status=500)


@cache_page(60 * 240)  # 30 minutos  
def obtener_datos_dashboard_reales():
    """Obtener datos reales para el dashboard desde la base de datos"""
    try:
        #  Métricas principales
        metricas_principales = calcular_metricas_principales()

        #  Análisis de precios (oportunidades de arbitraje)
        analisis_precios = calcular_oportunidades_precios()

        #  Análisis de volúmenes
        analisis_volumenes = calcular_analisis_volumenes()

        #  Oportunidades de mercado
        oportunidades_mercado = identificar_oportunidades_mercado()

        #  Análisis temporal
        analisis_temporal = calcular_analisis_temporal()

        return {
            'metricas_principales': metricas_principales,
            'analisis_precios': analisis_precios,
            'analisis_volumenes': analisis_volumenes,
            'oportunidades_mercado': oportunidades_mercado,
            'analisis_temporal': analisis_temporal
        }

    except Exception as e:
        logger.error(f"Error obteniendo datos del dashboard: {str(e)}")
        return {}

def calcular_metricas_principales():
    """Calcular métricas principales del dashboard"""
    try:
        # Total de registros
        total_registros = DatosComercializacion.objects.count()

        # Regiones activas
        regiones_activas = Region.objects.filter(
            datoscomercializacion__isnull=False
        ).distinct().count()

        # Productos únicos
        productos_unicos = Producto.objects.filter(
            datoscomercializacion__isnull=False
        ).distinct().count()

        # Mercados activos
        mercados_activos = Mercado.objects.filter(
            datoscomercializacion__isnull=False
        ).distinct().count()

        return {
            'margen_potencial_promedio': 35.2,  # Placeholder
            'productos_alto_margen': 8,         # Placeholder
            'regiones_activas': regiones_activas,
            'estacionalidad_promedio': 28.7     # Placeholder
        }

    except Exception as e:
        logger.error(f"Error calculando métricas principales: {str(e)}")
        return {
            'margen_potencial_promedio': 0,
            'productos_alto_margen': 0,
            'regiones_activas': 0,
            'estacionalidad_promedio': 0
        }

def calcular_oportunidades_precios():
    """Calcular oportunidades de arbitraje por diferencias de precios entre regiones"""
    try:
        # Obtener productos con mayor variación de precios
        productos_variacion = (DatosComercializacion.objects
                            .values('producto__nombre')
                            .annotate(
                                precio_max=Max('precio_promedio'),
                                precio_min=Min('precio_promedio'),
                                region_max=Max('region__nombre'),
                                region_min=Min('region__nombre')
                            )
                            .filter(precio_min__gt=0)
                            .annotate(
                                margen=((F('precio_max') - F('precio_min')) / F('precio_min') * 100)
                            )
                            .order_by('-margen')[:10])

        top_oportunidades = []
        for producto in productos_variacion:
            top_oportunidades.append({
                'producto': producto['producto__nombre'],
                'mejor_region': producto['region_max'],
                'peor_region': producto['region_min'],
                'precio_maximo': f"${producto['precio_max']:,.0f}",
                'precio_minimo': f"${producto['precio_min']:,.0f}",
                'diferencial_precio': f"${producto['precio_max'] - producto['precio_min']:,.0f}",
                'margen_potencial': round(producto['margen'], 1)
            })

        return {
            'top_oportunidades': top_oportunidades
        }

    except Exception as e:
        logger.error(f"Error calculando oportunidades de precios: {str(e)}")
        return {
            'top_oportunidades': []
        }

def calcular_analisis_volumenes():
    """Calcular análisis de volúmenes por región"""
    try:
        # Regiones con mayor volumen
        regiones_volumen = (DatosComercializacion.objects
                          .values('region__nombre')
                          .annotate(
                              volumen_total=Sum('volumen'),
                              productos_unicos=Count('producto', distinct=True),
                              mercados_activos=Count('mercado', distinct=True)
                          )
                          .order_by('-volumen_total')[:10])

        top_regiones = []
        for region in regiones_volumen:
            # Obtener producto más transado en la región
            producto_principal = (DatosComercializacion.objects
                                .filter(region__nombre=region['region__nombre'])
                                .values('producto__nombre')
                                .annotate(total_volumen=Sum('volumen'))
                                .order_by('-total_volumen')
                                .first())

            top_regiones.append({
                'region': region['region__nombre'],
                'producto_mas_transado': producto_principal['producto__nombre'] if producto_principal else 'N/A',
                'volumen_total': f"{region['volumen_total']:,.0f}",
                'mercados_activos': region['mercados_activos'],
                'liquidez': round(region['volumen_total'] / 100000, 1)  # Placeholder para liquidez
            })

        return {
            'top_regiones_volumen': top_regiones
        }

    except Exception as e:
        logger.error(f"Error calculando análisis de volúmenes: {str(e)}")
        return {
            'top_regiones_volumen': []
        }

def identificar_oportunidades_region(region_id, producto='', subsector='', año=''):
    """Oportunidades específicas para una región"""
    try:
        print(f" DEBUG OPORTUNIDADES REGIONAL - region_id: {region_id}")

        region = Region.objects.get(id_region=region_id)
        queryset = DatosComercializacion.objects.filter(region=region)

        #  AGREGAR FILTROS DE OUTLIERS AQUÍ
        queryset = queryset.filter(
            precio_promedio__gt=10,        # Mínimo $10
            precio_promedio__lt=50000,     # Máximo $50,000
            precio_minimo__gt=5,           # Mínimo $5
            precio_maximo__lt=100000,      # Máximo $100,000
            volumen__gt=0                  # Volumen positivo
        )

        print(f" DEBUG OPORTUNIDADES - Registros después de outliers: {queryset.count()}")

        if producto:
            queryset = queryset.filter(producto__nombre__icontains=producto)
        if subsector:
            queryset = queryset.filter(subsector__nombre__icontains=subsector)
        if año:
            try:
                queryset = queryset.filter(fecha__year=int(año))
            except ValueError:
                pass

        oportunidades = []

        # Productos con mayor margen en la región - CON FILTROS DE OUTLIERS
        productos_margen = (queryset
                          .values('producto__nombre')
                          .annotate(
                              margen=((Max('precio_promedio') - Min('precio_promedio')) / Min('precio_promedio') * 100),
                              total_registros=Count('id')
                          )
                          .filter(
                              margen__gt=10,           # Mínimo 10%
                              margen__lt=300,          #  MÁXIMO 300% (no millones)
                              total_registros__gt=1,
                              #  AGREGAR FILTROS DE PRECIO
                              precio_promedio__gt=10,
                              precio_promedio__lt=50000
                          )
                          .order_by('-margen')[:5])

        print(f" DEBUG OPORTUNIDADES - Productos con margen realista: {len(productos_margen)}")

        for producto_data in productos_margen:
            #  VERIFICAR QUE EL MARGEN SEA REAL
            if producto_data['margen'] <= 300:  # Máximo 300%
                potencial = 'Alto' if producto_data['margen'] > 100 else 'Medio' if producto_data['margen'] > 50 else 'Bajo'

                oportunidades.append({
                    'tipo': 'Arbitraje Local',
                    'descripcion': f"{producto_data['producto__nombre']} - Margen {producto_data['margen']:.1f}%",
                    'potencial': potencial,
                    'riesgo': 'Bajo',
                    'recomendacion': 'Explotar diferencia de precios entre mercados locales'
                })
                print(f" DEBUG OPORTUNIDADES - Margen REAL: {producto_data['producto__nombre']} - {producto_data['margen']:.1f}%")
            else:
                print(f" DEBUG OPORTUNIDADES - Margen IRREAL filtrado: {producto_data['producto__nombre']} - {producto_data['margen']:.1f}%")

        # Productos con alta demanda - CON FILTROS DE OUTLIERS
        productos_demanda = (queryset
                           .values('producto__nombre')
                           .annotate(
                               volumen_total=Sum('volumen'),
                               total_registros=Count('id')
                           )
                           .filter(
                               volumen_total__gt=100000,  # Volumen significativo
                               total_registros__gt=2
                           )
                           .order_by('-volumen_total')[:3])

        print(f" DEBUG OPORTUNIDADES - Productos con alta demanda: {len(productos_demanda)}")

        for producto_data in productos_demanda:
            volumen_formateado = f"{producto_data['volumen_total']:,.0f}" if producto_data['volumen_total'] else '0'
            oportunidades.append({
                'tipo': 'Mercado Estable',
                'descripcion': f"{producto_data['producto__nombre']} - Alto volumen ({volumen_formateado})",
                'potencial': 'Medio',
                'riesgo': 'Bajo',
                'recomendacion': 'Mercado consolidado con buena liquidez'
            })

        print(f" DEBUG OPORTUNIDADES - Oportunidades REALES totales: {len(oportunidades)}")

        return oportunidades

    except Region.DoesNotExist:
        print(f" DEBUG OPORTUNIDADES - Región {region_id} no existe")
        return []
    except Exception as e:
        print(f" DEBUG OPORTUNIDADES - Error identificando oportunidades regionales: {str(e)}")
        return []

@cache_page(60 * 240)  # 30 minutos  
def api_grafico_precios_temporales(request):
    """API para datos de gráfico de precios temporales"""
    try:
        print(f" DEBUG GRÁFICO PRECIOS - Iniciando")

        # Obtener parámetros de filtro
        region_id = request.GET.get('region_id', '')
        producto = request.GET.get('producto', '')
        subsector = request.GET.get('subsector', '')
        año = request.GET.get('año', '')

        print(f" DEBUG GRÁFICO PRECIOS - Filtros: region_id={region_id}, producto={producto}, subsector={subsector}, año={año}")

        queryset = DatosComercializacion.objects.all()

        # Aplicar filtros
        if region_id:
            try:
                region = Region.objects.get(id_region=int(region_id))
                queryset = queryset.filter(region=region)
                print(f" DEBUG GRÁFICO PRECIOS - Filtrado por región: {region.nombre}")
            except (ValueError, Region.DoesNotExist):
                print(f" DEBUG GRÁFICO PRECIOS - Región {region_id} no encontrada")

        if producto:
            queryset = queryset.filter(producto__nombre__icontains=producto)
        if subsector:
            queryset = queryset.filter(subsector__nombre__icontains=subsector)
        if año:
            try:
                queryset = queryset.filter(fecha__year=int(año))
            except ValueError:
                pass

        # Datos mensuales de precios
        datos_mensuales = (queryset
                          .values('fecha__year', 'fecha__month')
                          .annotate(
                              precio_promedio=Avg('precio_promedio'),
                              total_registros=Count('id')
                          )
                          .filter(total_registros__gt=0)
                          .order_by('fecha__year', 'fecha__month'))

        print(f" DEBUG GRÁFICO PRECIOS - Datos mensuales encontrados: {len(datos_mensuales)}")

        # Preparar datos para el gráfico
        labels = []
        precios = []

        for dato in datos_mensuales:
            fecha_str = f"{dato['fecha__month']}/{dato['fecha__year']}"
            labels.append(fecha_str)
            precio = float(dato['precio_promedio']) if dato['precio_promedio'] else 0
            precios.append(precio)
            print(f" DEBUG GRÁFICO PRECIOS - {fecha_str}: ${precio:.2f}")

        response_data = {
            'labels': labels,
            'datasets': [{
                'label': 'Precio Promedio',
                'data': precios,
                'borderColor': '#2e7d32',
                'backgroundColor': 'rgba(46, 125, 50, 0.1)',
                'tension': 0.4
            }]
        }

        print(f" DEBUG GRÁFICO PRECIOS - Datos enviados: {len(labels)} puntos")

        return JsonResponse(response_data)

    except Exception as e:
        print(f" DEBUG GRÁFICO PRECIOS - ERROR: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)

@cache_page(60 * 240)  # 30 minutos  
def api_grafico_distribucion_regiones(request):
    """API para gráfico de distribución por regiones"""
    try:
        print(f" DEBUG GRÁFICO DISTRIBUCIÓN - Iniciando")

        producto = request.GET.get('producto', '')
        subsector = request.GET.get('subsector', '')
        año = request.GET.get('año', '')
        region_id = request.GET.get('region_id', '')  # ← AGREGAR ESTE FILTRO

        print(f" DEBUG GRÁFICO DISTRIBUCIÓN - Filtros: region_id={region_id}, producto={producto}, subsector={subsector}, año={año}")

        queryset = DatosComercializacion.objects.all()

        # APLICAR FILTRO DE REGIÓN SI EXISTE
        if region_id:
            try:
                region = Region.objects.get(id_region=int(region_id))
                queryset = queryset.filter(region=region)
                print(f" DEBUG GRÁFICO DISTRIBUCIÓN - Filtrado por región: {region.nombre}")
            except (ValueError, Region.DoesNotExist):
                print(f" DEBUG GRÁFICO DISTRIBUCIÓN - Región {region_id} no encontrada")

        if producto:
            queryset = queryset.filter(producto__nombre__icontains=producto)
        if subsector:
            queryset = queryset.filter(subsector__nombre__icontains=subsector)
        if año:
            try:
                queryset = queryset.filter(fecha__year=int(año))
            except ValueError:
                pass

        # SI HAY FILTRO DE REGIÓN, mostrar mercados de esa región
        # SI NO HAY FILTRO DE REGIÓN, mostrar todas las regiones
        if region_id:
            # Mostrar MERCADOS de la región específica
            distribucion = (queryset
                          .values('mercado__nombre')
                          .annotate(
                              volumen_total=Sum('volumen'),
                              precio_promedio=Avg('precio_promedio'),
                              total_registros=Count('id')
                          )
                          .filter(volumen_total__isnull=False)
                          .order_by('-volumen_total')[:10])
            label_field = 'mercado__nombre'
            print(f" DEBUG GRÁFICO DISTRIBUCIÓN - Mostrando MERCADOS de la región {region_id}")
        else:
            # Mostrar REGIONES (vista nacional)
            distribucion = (queryset
                          .values('region__nombre')
                          .annotate(
                              volumen_total=Sum('volumen'),
                              precio_promedio=Avg('precio_promedio'),
                              total_registros=Count('id')
                          )
                          .filter(volumen_total__isnull=False)
                          .order_by('-volumen_total')[:10])
            label_field = 'region__nombre'
            print(f" DEBUG GRÁFICO DISTRIBUCIÓN - Mostrando REGIONES (vista nacional)")

        print(f" DEBUG GRÁFICO DISTRIBUCIÓN - Elementos con datos: {len(distribucion)}")

        labels = []
        volumenes = []
        precios = []

        for item in distribucion:
            labels.append(item[label_field])
            volumen = float(item['volumen_total']) if item['volumen_total'] else 0
            precio = float(item['precio_promedio']) if item['precio_promedio'] else 0
            volumenes.append(volumen)
            precios.append(precio)
            print(f" DEBUG GRÁFICO DISTRIBUCIÓN - {item[label_field]}: Volumen={volumen:,.0f}, Precio=${precio:.2f}")

        response_data = {
            'labels': labels,
            'datasets': [
                {
                    'label': 'Volumen Total',
                    'data': volumenes,
                    'backgroundColor': 'rgba(54, 162, 235, 0.6)',
                    'borderColor': 'rgba(54, 162, 235, 1)',
                    'borderWidth': 1
                },
                {
                    'label': 'Precio Promedio',
                    'data': precios,
                    'backgroundColor': 'rgba(255, 99, 132, 0.6)',
                    'borderColor': 'rgba(255, 99, 132, 1)',
                    'borderWidth': 1,
                    'type': 'line',
                    'yAxisID': 'y1'
                }
            ]
        }

        print(f" DEBUG GRÁFICO DISTRIBUCIÓN - Datos enviados: {len(labels)} elementos")

        return JsonResponse(response_data)

    except Exception as e:
        print(f" DEBUG GRÁFICO DISTRIBUCIÓN - ERROR: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)



@cache_page(60 * 240)  # 30 minutos  
def api_grafico_estacionalidad(request):
    """API para gráfico de estacionalidad por meses"""
    try:
        print(f" DEBUG GRÁFICO ESTACIONALIDAD - Iniciando")

        region_id = request.GET.get('region_id', '')
        producto = request.GET.get('producto', '')
        subsector = request.GET.get('subsector', '')

        print(f" DEBUG GRÁFICO ESTACIONALIDAD - Filtros: region_id={region_id}")

        queryset = DatosComercializacion.objects.all()

        #  FILTRAR OUTLIERS
        queryset = queryset.filter(
            precio_promedio__gt=10,
            precio_promedio__lt=50000,
            volumen__gt=0
        )

        if region_id:
            try:
                region = Region.objects.get(id_region=int(region_id))
                queryset = queryset.filter(region=region)
                print(f" DEBUG GRÁFICO ESTACIONALIDAD - Filtrado por región: {region.nombre}")
            except (ValueError, Region.DoesNotExist):
                print(f" DEBUG GRÁFICO ESTACIONALIDAD - Región {region_id} no encontrada")
        if producto:
            queryset = queryset.filter(producto__nombre__icontains=producto)
        if subsector:
            queryset = queryset.filter(subsector__nombre__icontains=subsector)

        total_registros = queryset.count()
        print(f" DEBUG GRÁFICO ESTACIONALIDAD - Total registros: {total_registros}")

        if total_registros == 0:
            print(" DEBUG GRÁFICO ESTACIONALIDAD - No hay datos")
            # Devolver datos vacíos pero con estructura correcta
            return JsonResponse({
                'labels': ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic'],
                'datasets': [
                    {
                        'label': 'Precio Promedio',
                        'data': [0]*12,
                        'borderColor': '#2e7d32',
                        'backgroundColor': 'rgba(46, 125, 50, 0.1)',
                        'yAxisID': 'y'
                    },
                    {
                        'label': 'Volumen Promedio',
                        'data': [0]*12,
                        'borderColor': '#1976d2',
                        'backgroundColor': 'rgba(25, 118, 210, 0.1)',
                        'yAxisID': 'y1'
                    }
                ]
            })

        # Estacionalidad por mes (promedio de todos los años)
        estacionalidad = (queryset
                         .values('fecha__month')
                         .annotate(
                             precio_promedio=Avg('precio_promedio'),
                             volumen_promedio=Avg('volumen'),
                             total_registros=Count('id')
                         )
                         .filter(total_registros__gt=0)
                         .order_by('fecha__month'))

        print(f" DEBUG GRÁFICO ESTACIONALIDAD - Meses con datos: {len(estacionalidad)}")

        meses = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun',
                'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']

        precios = [0] * 12
        volumenes = [0] * 12

        for dato in estacionalidad:
            mes_index = dato['fecha__month'] - 1
            if 0 <= mes_index < 12:
                precio = float(dato['precio_promedio']) if dato['precio_promedio'] else 0
                volumen = float(dato['volumen_promedio']) if dato['volumen_promedio'] else 0
                precios[mes_index] = precio
                volumenes[mes_index] = volumen
                print(f" DEBUG GRÁFICO ESTACIONALIDAD - Mes {mes_index+1}: Precio=${precio:.2f}, Volumen={volumen:,.0f}")

        response_data = {
            'labels': meses,
            'datasets': [
                {
                    'label': 'Precio Promedio',
                    'data': precios,
                    'borderColor': '#2e7d32',
                    'backgroundColor': 'rgba(46, 125, 50, 0.1)',
                    'yAxisID': 'y'
                },
                {
                    'label': 'Volumen Promedio',
                    'data': volumenes,
                    'borderColor': '#1976d2',
                    'backgroundColor': 'rgba(25, 118, 210, 0.1)',
                    'yAxisID': 'y1'
                }
            ]
        }

        print(f" DEBUG GRÁFICO ESTACIONALIDAD - Datos enviados: 12 meses")

        return JsonResponse(response_data)

    except Exception as e:
        print(f" DEBUG GRÁFICO ESTACIONALIDAD - ERROR: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)



# Funciones placeholder para otros tipos de análisis (las implementaremos después)
def calcular_analisis_precios_producto(producto, subsector='', año=''):
    """Análisis de precios para un producto específico"""
    return calcular_analisis_precios_filtrados(subsector, producto, año, '')

def calcular_analisis_volumenes_producto(producto, subsector='', año=''):
    """Análisis de volúmenes para un producto específico"""
    return calcular_analisis_volumenes_filtrados(subsector, producto, año, '')

def identificar_oportunidades_producto(producto, subsector='', año=''):
    """Oportunidades para un producto específico"""
    return identificar_oportunidades_mercado_filtradas(subsector, producto, año, '')

def calcular_analisis_precios_subsector(subsector, año=''):
    """Análisis de precios para un subsector"""
    return calcular_analisis_precios_filtrados(subsector, '', año, '')

def calcular_analisis_volumenes_subsector(subsector, año=''):
    """Análisis de volúmenes para un subsector"""
    return calcular_analisis_volumenes_filtrados(subsector, '', año, '')

def identificar_oportunidades_subsector(subsector, año=''):
    """Oportunidades para un subsector"""
    return identificar_oportunidades_mercado_filtradas(subsector, '', año, '')


def identificar_oportunidades_mercado():
    """Identificar oportunidades de mercado basadas en datos reales"""
    try:
        # Analizar productos con mayor margen
        productos_margen = (DatosComercializacion.objects
                          .values('producto__nombre')
                          .annotate(
                              margen=((Max('precio_promedio') - Min('precio_promedio')) / Min('precio_promedio') * 100)
                          )
                          .filter(margen__gt=30)
                          .order_by('-margen')[:5])

        oportunidades = []

        for producto in productos_margen:
            oportunidades.append({
                'tipo': 'Arbitraje Regional',
                'descripcion': f"{producto['producto__nombre']} - Margen {producto['margen']:.1f}%",
                'potencial': 'Alto' if producto['margen'] > 50 else 'Medio',
                'riesgo': 'Medio',
                'recomendacion': f"Explotar diferencia de precios entre regiones para {producto['producto__nombre']}"
            })

        # Agregar oportunidades basadas en volumen
        productos_volumen = (DatosComercializacion.objects
                           .values('producto__nombre')
                           .annotate(volumen_total=Sum('volumen'))
                           .order_by('-volumen_total')[:3])

        for producto in productos_volumen:
            oportunidades.append({
                'tipo': 'Mercado Líquido',
                'descripcion': f"{producto['producto__nombre']} - Alto volumen",
                'potencial': 'Medio',
                'riesgo': 'Bajo',
                'recomendacion': f"Mercado estable para {producto['producto__nombre']} con buena liquidez"
            })

        return oportunidades

    except Exception as e:
        logger.error(f"Error identificando oportunidades: {str(e)}")
        return []

def calcular_analisis_temporal():
    """Calcular análisis temporal y estacionalidad"""
    try:
        # Análisis temporal básico
        return {
            'estacionalidad_productos': [
                {
                    'producto': 'Producto Principal',
                    'mejor_mes_comprar': 'Enero',
                    'mejor_mes_vender': 'Junio',
                    'variacion_estacional': 25.5,
                    'tendencia': 'Alcista'
                }
            ]
        }

    except Exception as e:
        logger.error(f"Error calculando análisis temporal: {str(e)}")
        return {
            'estacionalidad_productos': []
        }